# Consider a source workbook, say source.xlsx and destination workbook, say destination.xlsx. The latter file is empty
# to where the contents of the former file will be copied. Below is the example of the former file (source.xlsx).
from openpyxl import load_workbook

# Load the workbooks
src_wb = load_workbook('source.xlsx')
des_wb = load_workbook('destination.xlsx')

# Read the sheets to be copied
src_sheet = src_wb['Sheet1']
des_sheet = des_wb['Sheet1']

# Copy all the rows and columns
for i in range(1, src_sheet.max_row+1):
    for j in range(1, src_sheet.max_column+1):
        des_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

# Save the workbooks
src_wb.save('source.xlsx')
des_wb.save('destination.xlsx')
